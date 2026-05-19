from django.shortcuts import render

import sys
import os
import json
import pandas as pd
import re
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from predictions.excel_processor import process_uploaded_file, generate_template
from ml.predictor import predict_from_dataframe
from ml.preprocessing import preprocess_uploaded_file
from ml.evaluate_model import get_metrics_dict
from ml.explainability import generate_explanation
from .models import PredictionBatch, CustomerPrediction


# ─────────────────────────────────────────────────────────────
# 1. UPLOAD & PREDICT
# POST /api/upload/
# ─────────────────────────────────────────────────────────────
class UploadPredictView(APIView):

    def post(self, request):
        # Step 1 — Check file was provided
        if 'file' not in request.FILES:
            return Response(
                {"error": "No file provided. Please upload an Excel or CSV file."},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES['file']

        # Step 2 — Process and validate file
        result = process_uploaded_file(file)
        if not result["success"]:
            return Response(
                {"error": result["error"], **{k: v for k, v in result.items() if k not in ["success", "error"]}},
                status=status.HTTP_400_BAD_REQUEST
            )

        df = result["dataframe"]
        
        PredictionBatch.objects.filter(file_name=file.name).delete()

        # Step 3 — Create batch record in database
        batch = PredictionBatch.objects.create(
            file_name=file.name,
            total_customers=len(df),
            status='processing'
        )

        try:
            # Step 4 — Preprocess for model
            X_scaled = preprocess_uploaded_file(df)

            # Step 5 — Run predictions
            predictions = predict_from_dataframe(X_scaled)

            # Step 6 — Count risk levels
            low_count    = sum(1 for p in predictions if p["risk_level"] == "Low Risk")
            medium_count = sum(1 for p in predictions if p["risk_level"] == "Medium Risk")
            high_count   = sum(1 for p in predictions if p["risk_level"] == "High Risk")

            # Step 7 — Save each customer prediction to database
            customer_objects = []
            for i, pred in enumerate(predictions):
                row = df.iloc[i]
                user_email = str(row.get("email", ""))
                user_contact = str(row.get("contact_number", "")).strip()
                val = str(row.get("employment", row.get("employment_years", "0"))).strip()
                match = re.search(r'\d+', val)
                clean_employment = match.group(0) if match else "0"
                customer_name = str(row.get("name", f"Customer {i+1}"))
                customer_objects.append(CustomerPrediction(
                    batch               = batch,
                    row_index           = i,
                    name                = customer_name,
                    email               = user_email,
                    contact_number      = user_contact,
                    explanation = str(row.get("name", f"Customer {i+1}")), 
                    checking_status     = str(row.get("checking_status", "")),
                    duration            = row.get("duration", None),
                    credit_history      = str(row.get("credit_history", "")),
                    purpose             = str(row.get("purpose", "")),
                    credit_amount       = row.get("credit_amount", None),
                    savings_status      = str(row.get("savings_status", "")),
                    employment          = clean_employment,
                    installment_commitment = row.get("installment_commitment", None),
                    personal_status     = str(row.get("personal_status", "")),
                    other_parties       = str(row.get("other_parties", "")),
                    residence_since     = row.get("residence_since", None),
                    property_magnitude  = str(row.get("property_magnitude", "")),
                    age                 = row.get("age", None),
                    other_payment_plans = str(row.get("other_payment_plans", "")),
                    housing             = str(row.get("housing", "")),
                    existing_credits    = row.get("existing_credits", None),
                    job                 = str(row.get("job", "")),
                    num_dependents      = row.get("num_dependents", None),
                    own_telephone       = str(row.get("own_telephone", "")),
                    foreign_worker      = str(row.get("foreign_worker", "")),
                    raw_prediction      = pred["raw_prediction"],
                    risk_level          = pred["risk_level"],
                    risk_color          = pred["risk_color"],
                    confidence          = pred["confidence"],
                    prob_good           = pred["probabilities"].get("good", 0),
                    prob_bad            = pred["probabilities"].get("bad", 0),
                ))

            # Bulk save for performance
            CustomerPrediction.objects.bulk_create(customer_objects)

            # Step 8 — Update batch with final counts
            batch.low_risk_count    = low_count
            batch.medium_risk_count = medium_count
            batch.high_risk_count   = high_count
            batch.status            = 'done'
            batch.save()

            # Step 9 — Return response
            return Response({
                "success":    True,
                "batch_id":   batch.id,
                "file_name":  file.name,
                "total":      len(predictions),
                "summary": {
                    "low_risk":    low_count,
                    "medium_risk": medium_count,
                    "high_risk":   high_count,
                },
                "predictions": [
                    {
                        "row":             p["row_index"] + 1,
                        "name": str(df.iloc[p["row_index"]].get("name", "")),
                        "risk_level":      p["risk_level"],
                        "risk_color":      p["risk_color"],
                        "confidence":      p["confidence"],
                        "raw_prediction":  p["raw_prediction"],
                        "probabilities":   p["probabilities"],
                        # Include key customer fields for display
                        "age":            float(df.iloc[p["row_index"]].get("age", 0)),
                        "credit_amount":  float(df.iloc[p["row_index"]].get("credit_amount", 0)),
                        "duration":       float(df.iloc[p["row_index"]].get("duration", 0)),
                        "purpose":        str(df.iloc[p["row_index"]].get("purpose", "")),
                    }
                    
                    for p in predictions
                ]
            }, status=status.HTTP_200_OK)

        except Exception as e:
            batch.status = 'failed'
            batch.save()
            return Response(
                {"error": f"Prediction failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ─────────────────────────────────────────────────────────────
# 2. LIST ALL BATCHES (for Reports page)
# GET /api/batches/
# ─────────────────────────────────────────────────────────────
class BatchListView(APIView):

    def get(self, request):
        batches = PredictionBatch.objects.all()
        data = [
            {
                "id":               b.id,
                "file_name":        b.file_name,
                "uploaded_at":      b.uploaded_at.strftime("%Y-%m-%d %H:%M"),
                "total_customers":  b.total_customers,
                "low_risk":         b.low_risk_count,
                "medium_risk":      b.medium_risk_count,
                "high_risk":        b.high_risk_count,
                "status":           b.status,
            }
            for b in batches
        ]
        return Response(data, status=status.HTTP_200_OK)


# ─────────────────────────────────────────────────────────────
# 3. GET SINGLE BATCH DETAILS
# GET /api/batches/<batch_id>/
# ─────────────────────────────────────────────────────────────
class BatchDetailView(APIView):

    def get(self, request, batch_id):
        try:
            batch = PredictionBatch.objects.get(id=batch_id)
        except PredictionBatch.DoesNotExist:
            return Response({"error": "Batch not found."}, status=status.HTTP_404_NOT_FOUND)

        customers = CustomerPrediction.objects.filter(batch=batch)
        data = {
            "id":              batch.id,
            "file_name":       batch.file_name,
            "uploaded_at":     batch.uploaded_at.strftime("%Y-%m-%d %H:%M"),
            "total_customers": batch.total_customers,
            "summary": {
                "low_risk":    batch.low_risk_count,
                "medium_risk": batch.medium_risk_count,
                "high_risk":   batch.high_risk_count,
            },
            "customers": [
                {
                    "row":            c.row_index + 1,
                    "name":           c.name,
                    "email":          c.email,
                    "contact_number": c.contact_number,
                    "age":            c.age,
                    "credit_amount":  c.credit_amount,
                    "duration":       c.duration,
                    "purpose":        c.purpose,
                    "employment":     c.employment,
                    "housing":        c.housing,
                    "risk_level":     c.risk_level,
                    "risk_color":     c.risk_color,
                    "confidence":     c.confidence,
                    "raw_prediction": c.raw_prediction,
                    "probabilities": {
                        "good": c.prob_good,
                        "bad":  c.prob_bad,
                    },
                }
                for c in customers
            ]
        }
        return Response(data, status=status.HTTP_200_OK)


# ─────────────────────────────────────────────────────────────
# 4. DOWNLOAD REPORT AS EXCEL
# GET /api/batches/<batch_id>/download/
# ─────────────────────────────────────────────────────────────
# class DownloadReportView(APIView):

#     def get(self, request, batch_id):
#         try:
#             batch = PredictionBatch.objects.get(id=batch_id)
#         except PredictionBatch.DoesNotExist:
#             return Response({"error": "Batch not found."}, status=status.HTTP_404_NOT_FOUND)

#         customers = CustomerPrediction.objects.filter(batch=batch),
        

#         # Build Excel workbook
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Predictions"

#         # Header row
#         headers = [
#             "Name", "Email", "Contact", "Age", "Credit Amount", "Duration", "Purpose",
#             "Employment", "Housing", "Risk Level", "Confidence (%)",
#             "Prob Good (%)", "Prob Bad (%)"
#         ]
#         ws.append(headers)

#         # Style header row
#         header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
#         for cell in ws[1]:
#             cell.fill      = header_fill
#             cell.font      = Font(color="FFFFFF", bold=True)
#             cell.alignment = Alignment(horizontal="center")

#         # Color mapping for risk levels
#         risk_colors = {
#             "Low Risk":    "C6EFCE",  # green
#             "Medium Risk": "FFEB9C",  # yellow
#             "High Risk":   "FFC7CE",  # red
#         }

#         # Data rows
#         for c in customers:
#             row = [
#                 c.row_index + 1,
#                 c.age,
#                 c.credit_amount,
#                 c.duration,
#                 c.purpose,
#                 c.employment,
#                 c.housing,
#                 c.risk_level,
#                 c.confidence,
#                 c.prob_good,
#                 c.prob_bad,
#             ]
#             ws.append(row)

#             # Color the Risk Level cell
#             risk_cell = ws.cell(row=ws.max_row, column=8)
#             color = risk_colors.get(c.risk_level, "FFFFFF")
#             risk_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
#             risk_cell.font = Font(bold=True)
#             risk_cell.alignment = Alignment(horizontal="center")

#         # Auto-fit column widths
#         for col in ws.columns:
#             max_len = max(len(str(cell.value or "")) for cell in col)
#             ws.column_dimensions[col[0].column_letter].width = max_len + 4

#         # Return as downloadable file
#         response = HttpResponse(
#             content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
#         response["Content-Disposition"] = f'attachment; filename="predictions_batch_{batch_id}.xlsx"'
#         wb.save(response)
#         return response


class DownloadReportView(APIView):
    def get(self, request, batch_id):
        try:
            batch = PredictionBatch.objects.get(id=batch_id)
        except PredictionBatch.DoesNotExist:
            return Response({"error": "Batch not found."}, status=status.HTTP_404_NOT_FOUND)

        # UPDATED: Removed the comma at the end and added .order_by
        customers = CustomerPrediction.objects.filter(batch=batch).order_by('row_index')

        # Build Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Predictions"

        # Header row (Matches your request: No "Row" column, starts with Name)
        headers = [
            "Name", "Email", "Contact", "Age", "Credit Amount", "Duration", "Purpose",
            "Employment", "Housing", "Risk Level", "Confidence (%)",
            "Prob Good (%)", "Prob Bad (%)"
        ]
        ws.append(headers)

        # Style header row
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Color mapping for risk levels
        risk_colors = {
            "Low Risk":    "C6EFCE",  # green
            "Medium Risk": "FFEB9C",  # yellow
            "High Risk":   "FFC7CE",  # red
        }

        # Data rows
        for c in customers:
            # UPDATED: Row list now perfectly matches the 'headers' list order
            row = [
                c.name or f"Customer {c.row_index + 1}",
                c.email or "",
                c.contact_number or "",
                c.age,
                c.credit_amount,
                c.duration,
                c.purpose,
                c.employment,
                c.housing,
                c.risk_level,
                f"{c.confidence}%",
                c.prob_good,
                c.prob_bad,
            ]
            ws.append(row)

            # Color the Risk Level cell (Now Column 10 because Name/Email/Contact shifted it)
            risk_cell = ws.cell(row=ws.max_row, column=10) 
            color = risk_colors.get(c.risk_level, "FFFFFF")
            risk_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            risk_cell.font = Font(bold=True)
            risk_cell.alignment = Alignment(horizontal="center")

            # Force Contact Number column (Column 3) to be text to keep leading zeros
            ws.cell(row=ws.max_row, column=3).number_format = '@'

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        # Return as downloadable file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="predictions_batch_{batch_id}.xlsx"'
        wb.save(response)
        return response


# ─────────────────────────────────────────────────────────────
# 5. DOWNLOAD BLANK TEMPLATE
# GET /api/template/
# ─────────────────────────────────────────────────────────────
class DownloadTemplateView(APIView):

    def get(self, request):
        df = generate_template()
        
        
        df.insert(1, "email", "customer@example.com")
        df.insert(2, "contact_number", "+923182538281")

        wb = Workbook()
        ws = wb.active
        ws.title = "Customer Data"

        # Write headers
        ws.append(list(df.columns))
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Write example row
        ws.append(list(df.iloc[0]))
        
        
        for cell in ws['C']:
            cell.number_format = '@'

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="credit_risk_template.xlsx"'
        wb.save(response)
        return response


# ─────────────────────────────────────────────────────────────
# 6. MODEL METRICS
# GET /api/metrics/
# ─────────────────────────────────────────────────────────────
class ModelMetricsView(APIView):

    def get(self, request):
        metrics = get_metrics_dict()
        if not metrics:
            return Response(
                {"error": "Model metrics not found. Run train_model.py first."},
                status=status.HTTP_404_NOT_FOUND
            )
        return Response({
            "accuracy":   metrics.get("accuracy"),
            "precision":  metrics.get("precision"),
            "recall":     metrics.get("recall"),
            "f1_score":   metrics.get("f1_score"),
            "cv_mean":    metrics.get("cv_mean"),
            "cv_std":     metrics.get("cv_std"),
        }, status=status.HTTP_200_OK)
        
        

def format_employment_simple(raw_value):
    """Extracts just the leading number from the employment category."""
    val = str(raw_value).strip()
    
    if val == "<1 yr":
        return "0"  # or "1" depending on how you want to represent it
    
    # This regex finds the first digit in strings like "1<=X<4 yrs" or ">=7 yrs"
    match = re.search(r'\d+', val)
    return match.group(0) if match else "0"


class ExplainView(APIView):

    def get(self, request, batch_id, row_index):
        try:
            batch    = PredictionBatch.objects.get(id=batch_id)
            customer = CustomerPrediction.objects.get(batch=batch, row_index=row_index)
        except (PredictionBatch.DoesNotExist, CustomerPrediction.DoesNotExist):
            return Response({"error": "Customer not found."}, status=status.HTTP_404_NOT_FOUND)

        # Build customer data dict for explainability
        customer_data = {
            "checking_status":        customer.checking_status,
            "duration":               customer.duration,
            "credit_history":         customer.credit_history,
            "purpose":                customer.purpose,
            "credit_amount":          customer.credit_amount,
            "savings_status":         customer.savings_status,
            "employment":             customer.employment,
            "installment_commitment": customer.installment_commitment,
            "personal_status":        customer.personal_status,
            "other_parties":          customer.other_parties,
            "residence_since":        customer.residence_since,
            "property_magnitude":     customer.property_magnitude,
            "age":                    customer.age,
            "other_payment_plans":    customer.other_payment_plans,
            "housing":                customer.housing,
            "existing_credits":       customer.existing_credits,
            "job":                    customer.job,
            "num_dependents":         customer.num_dependents,
            "own_telephone":          customer.own_telephone,
            "foreign_worker":         customer.foreign_worker,
        }

        prediction = {
            "risk_level":      customer.risk_level,
            "confidence":      customer.confidence,
            "raw_prediction":  customer.raw_prediction,
        }

        explanation = generate_explanation(customer_data, prediction)

        return Response({
            "batch_id":   batch_id,
            "row_index":  row_index,
            "name":       customer.name,
            "age":        customer.age,
            "credit_amount": customer.credit_amount,
            "purpose":    customer.purpose,
            **explanation,
        }, status=status.HTTP_200_OK)